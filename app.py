import pymysql
pymysql.install_as_MySQLdb()
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, g
from werkzeug.security import generate_password_hash, check_password_hash
import sqlite3
from datetime import datetime, date, timedelta
import os
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from io import BytesIO
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import send_file
from utils.email_alerts import check_and_send_expiry_alerts, send_test_email

def add_sales_extra_columns():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    try:
        cursor.execute("ALTER TABLE sales ADD COLUMN discount_amount REAL DEFAULT 0")
    except sqlite3.OperationalError:
        pass  # column already exists

    try:
        cursor.execute("ALTER TABLE sales ADD COLUMN tax_amount REAL DEFAULT 0")
    except sqlite3.OperationalError:
        pass

    try:
        cursor.execute("ALTER TABLE sales ADD COLUMN gross_amount REAL DEFAULT 0")
    except sqlite3.OperationalError:
        pass

    conn.commit()
    conn.close()

app = Flask(__name__)
app.secret_key = 'pharmacy-secret-key-2025-change-in-production'

# Database path
DB_PATH = os.path.join('database', 'pharmacy.db')

def get_db():
    """Get database connection with row factory"""
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(error):
    """Close database connection"""
    if hasattr(g, 'db'):
        g.db.close()

def init_db():
    """Initialize database with all tables"""
    if not os.path.exists('database'):
        os.makedirs('database')
    
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Suppliers table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS suppliers (
            supplier_id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_name TEXT NOT NULL,
            contact_number TEXT NOT NULL,
            email TEXT,
            address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Medicines table - FIXED: DATE -> TEXT
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS medicines (
            medicine_id INTEGER PRIMARY KEY AUTOINCREMENT,
            medicine_name TEXT NOT NULL,
            category TEXT NOT NULL,
            batch_number TEXT NOT NULL,
            price REAL NOT NULL,
            quantity INTEGER NOT NULL DEFAULT 0,
            expiry_date TEXT NOT NULL,
            supplier_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (supplier_id) REFERENCES suppliers(supplier_id)
        )
    ''')
    
    # Sales table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sales (
            sale_id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            customer_name TEXT,
            total_amount REAL NOT NULL,
            payment_method TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Sale items table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sale_items (
            item_id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER NOT NULL,
            medicine_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL,
            price REAL NOT NULL,
            subtotal REAL NOT NULL,
            FOREIGN KEY (sale_id) REFERENCES sales(sale_id),
            FOREIGN KEY (medicine_id) REFERENCES medicines(medicine_id)
        )
    ''')
    
    # Create default admin user
    cursor.execute('SELECT * FROM users WHERE username = ?', ('admin',))
    if not cursor.fetchone():
        hashed_password = generate_password_hash('admin123')
        cursor.execute('INSERT INTO users (username, password, role) VALUES (?, ?, ?)',
                      ('admin', hashed_password, 'admin'))
    
    conn.commit()
    conn.close()
    print("✅ Database initialized successfully!")

# Login required decorator
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login first', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Routes
@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        conn.close()
        
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['user_id']
            session['username'] = user['username']
            session['role'] = user['role']
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully', 'success')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()
    
    total_medicines = conn.execute('SELECT COUNT(*) as count FROM medicines').fetchone()['count']
    low_stock = conn.execute('SELECT COUNT(*) as count FROM medicines WHERE quantity < 10').fetchone()['count']
    
    today = date.today()
    today_sales = conn.execute(
        'SELECT COALESCE(SUM(total_amount), 0) as total FROM sales WHERE DATE(sale_date) = ?',
        (today,)
    ).fetchone()['total']
    
    expired = conn.execute(
        'SELECT COUNT(*) as count FROM medicines WHERE DATE(expiry_date) < DATE("now")'
    ).fetchone()['count']
    
    near_expiry = conn.execute(
        'SELECT COUNT(*) as count FROM medicines WHERE DATE(expiry_date) BETWEEN DATE("now") AND DATE("now", "+30 days")'
    ).fetchone()['count']
    
    conn.close()
    
    return render_template('dashboard.html',
                         total_medicines=total_medicines,
                         low_stock=low_stock,
                         today_sales=today_sales,
                         expired=expired,
                         near_expiry=near_expiry)



@app.route('/send-expiry-email-alerts', methods=['POST'])
@login_required
def send_expiry_email_alerts():
    try:
        success = check_and_send_expiry_alerts()
        
        return jsonify({
            'success': success,
            'message': 'Email alerts sent to admin!' if success else 'Failed to send alerts'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

@app.route('/test-email', methods=['POST'])
@login_required
def test_email():
    try:
        success = send_test_email()
        
        return jsonify({
            'success': success,
            'message': 'Test email sent!' if success else 'Failed to send test email'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500
@app.route('/medicines')
@login_required
def medicines():
    conn = get_db()
    medicines_raw = conn.execute('''
        SELECT m.*, s.supplier_name 
        FROM medicines m
        LEFT JOIN suppliers s ON m.supplier_id = s.supplier_id
        ORDER BY m.medicine_name
    ''').fetchall()
    
    suppliers_raw = conn.execute('SELECT * FROM suppliers ORDER BY supplier_name').fetchall()
    conn.close()
    
    medicines = [dict(row) for row in medicines_raw]
    suppliers = [dict(row) for row in suppliers_raw]
    
    return render_template('medicine_management.html', medicines=medicines, suppliers=suppliers)

@app.route('/medicines/add', methods=['POST'])
@login_required
def add_medicine():
    try:
        medicine_name = request.form['medicine_name']
        category = request.form['category']
        batch_number = request.form['batch_number']
        price = float(request.form['price'])
        quantity = int(request.form['quantity'])
        expiry_date = request.form['expiry_date']
        supplier_id = request.form.get('supplier_id') or None
        
        conn = get_db()
        conn.execute('''
            INSERT INTO medicines (medicine_name, category, batch_number, price, quantity, expiry_date, supplier_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (medicine_name, category, batch_number, price, quantity, expiry_date, supplier_id))
        conn.commit()
        conn.close()
        
        flash('Medicine added successfully!', 'success')
    except Exception as e:
        flash(f'Error adding medicine: {str(e)}', 'error')
    
    return redirect(url_for('medicines'))
@app.route('/medicines/update/<int:id>', methods=['POST'])
@login_required
def update_medicine(id):
    try:
        new_quantity = int(request.form['quantity'])
        
        conn = get_db()
        conn.execute('''
            UPDATE medicines 
            SET medicine_name=?, category=?, batch_number=?, price=?, 
                quantity=?, expiry_date=?, supplier_id=?
            WHERE medicine_id=?
        ''', (
            request.form['medicine_name'],
            request.form['category'],
            request.form['batch_number'],
            float(request.form['price']),
            new_quantity,
            request.form['expiry_date'],
            request.form.get('supplier_id'),
            id
        ))
        
        # ← ADD THIS: Check if updated quantity is below 5
        if new_quantity < 5 and new_quantity > 0:
            medicine = conn.execute('''
                SELECT m.*, s.supplier_name 
                FROM medicines m
                LEFT JOIN suppliers s ON m.supplier_id = s.supplier_id
                WHERE m.medicine_id = ?
            ''', (id,)).fetchone()
            
            if medicine:
                from utils.email_alerts import send_low_stock_alert
                import threading
                email_thread = threading.Thread(
                    target=send_low_stock_alert,
                    args=([dict(medicine)],)
                )
                email_thread.daemon = True
                email_thread.start()
        
        conn.commit()
        conn.close()
        
        flash('Medicine updated successfully!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('medicines'))

@app.route('/medicines/delete/<int:id>')
@login_required
def delete_medicine(id):
    try:
        conn = get_db()
        conn.execute('DELETE FROM medicines WHERE medicine_id = ?', (id,))
        conn.commit()
        conn.close()
        flash('Medicine deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting medicine: {str(e)}', 'error')
    
    return redirect(url_for('medicines'))

@app.route('/stock')
@login_required
def stock():
    conn = get_db()
    
    medicines = conn.execute('''
        SELECT m.*, s.supplier_name,
        CASE 
            WHEN m.quantity = 0 THEN 'Out of Stock'
            WHEN m.quantity < 10 THEN 'Low Stock'
            ELSE 'In Stock'
        END as stock_status
        FROM medicines m
        LEFT JOIN suppliers s ON m.supplier_id = s.supplier_id
        ORDER BY m.quantity ASC
    ''').fetchall()
    
    conn.close()
    return render_template('stock_monitoring.html', medicines=[dict(row) for row in medicines])

@app.route('/expiry')
@login_required  # FIXED: Added login_required
def expiry():
    conn = get_db()
    cursor = conn.cursor()
    today = date.today()

    # Expired medicines
    cursor.execute("""
        SELECT * FROM medicines
        WHERE DATE(expiry_date) < DATE('now')
        ORDER BY expiry_date ASC
    """)
    expired_raw = cursor.fetchall()
    
    expired = []
    for row in expired_raw:
        med = dict(row)
        exp_date = datetime.strptime(med['expiry_date'], '%Y-%m-%d').date()
        med['expiry_date_fmt'] = exp_date.strftime('%d-%m-%Y')
        med['days_expired'] = (today - exp_date).days
        expired.append(med)

    # Near expiry (next 30 days)
    cursor.execute("""
        SELECT * FROM medicines
        WHERE DATE(expiry_date) BETWEEN DATE('now') AND DATE('now', '+30 days')
        ORDER BY expiry_date ASC
    """)
    near_expiry_raw = cursor.fetchall()
    
    near_expiry = []
    for row in near_expiry_raw:
        med = dict(row)
        exp_date = datetime.strptime(med['expiry_date'], '%Y-%m-%d').date()
        med['expiry_date_fmt'] = exp_date.strftime('%d-%m-%Y')
        med['days_left'] = (exp_date - today).days
        near_expiry.append(med)

    conn.close()
    return render_template('expiry_alert.html', expired=expired, near_expiry=near_expiry)
@app.route('/billing')
@login_required
def billing():
    conn = get_db()
    # Get only non-expired medicines with stock
    medicines = conn.execute('''
        SELECT * FROM medicines 
        WHERE quantity > 0 AND expiry_date >= DATE("now")
        ORDER BY medicine_name
    ''').fetchall()
    conn.close()
    
    return render_template('billing.html', medicines=medicines)

@app.route('/billing/create', methods=['POST'])
@login_required
def create_bill():
    try:
        customer_name = request.form.get('customer_name', '')
        payment_method = request.form['payment_method']
        items = request.form.getlist('medicine_id[]')
        quantities = request.form.getlist('quantity[]')
        
        if not items:
            return jsonify({'success': False, 'message': 'No items in cart'}), 400
        
        conn = get_db()
        total_amount = 0
        low_stock_medicines = []  # ← ADD THIS
        
        cursor = conn.execute(
            'INSERT INTO sales (customer_name, total_amount, payment_method) VALUES (?, ?, ?)',
            (customer_name, 0, payment_method)
        )
        sale_id = cursor.lastrowid
        
        for i, medicine_id in enumerate(items):
            quantity = int(quantities[i])
            medicine = conn.execute('SELECT * FROM medicines WHERE medicine_id = ?', (medicine_id,)).fetchone()
            
            if not medicine:
                conn.rollback()
                conn.close()
                return jsonify({'success': False, 'message': f'Medicine not found'}), 400
            
            if medicine['quantity'] < quantity:
                conn.rollback()
                conn.close()
                return jsonify({'success': False, 'message': f'Insufficient stock'}), 400
            
            subtotal = medicine['price'] * quantity
            total_amount += subtotal
            
            conn.execute(
                'INSERT INTO sale_items (sale_id, medicine_id, quantity, price, subtotal) VALUES (?, ?, ?, ?, ?)',
                (sale_id, medicine_id, quantity, medicine['price'], subtotal)
            )
            
            # Update stock
            new_quantity = medicine['quantity'] - quantity
            conn.execute(
                'UPDATE medicines SET quantity = ? WHERE medicine_id = ?',
                (new_quantity, medicine_id)
            )
            
            # ← ADD THIS: Check if stock went below 5
            if new_quantity < 5 and new_quantity > 0:
                low_stock_medicines.append({
                    'medicine_name': medicine['medicine_name'],
                    'category': medicine['category'],
                    'quantity': new_quantity
                })
        
        conn.execute('UPDATE sales SET total_amount = ? WHERE sale_id = ?', (total_amount, sale_id))
        conn.commit()
        conn.close()
        
        # ← ADD THIS: Send automatic email if any medicine is now low stock
        if low_stock_medicines:
            from utils.email_alerts import send_low_stock_alert
            import threading
            # Send email in background thread so billing doesn't wait
            email_thread = threading.Thread(
                target=send_low_stock_alert, 
                args=(low_stock_medicines,)
            )
            email_thread.daemon = True
            email_thread.start()
        
        return jsonify({
            'success': True,
            'sale_id': sale_id,
            'total_amount': total_amount,
            'message': 'Bill created successfully'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500
@app.route('/api/search-medicine')
@login_required
def search_medicine_api():
    query = request.args.get('q', '')
    conn = get_db()
    
    medicines = conn.execute('''
        SELECT medicine_id, medicine_name, category, batch_number, price, quantity, expiry_date
        FROM medicines 
        WHERE (medicine_name LIKE ? OR batch_number LIKE ?)
        AND quantity > 0 
        AND expiry_date >= DATE("now")
        ORDER BY medicine_name
        LIMIT 10
    ''', (f'%{query}%', f'%{query}%')).fetchall()
    
    conn.close()
    
    return jsonify([dict(med) for med in medicines])
    
@app.route('/suppliers')
@login_required
def suppliers():
    conn = get_db()
    suppliers_raw = conn.execute('SELECT * FROM suppliers ORDER BY supplier_name').fetchall()
    conn.close()
    suppliers = [dict(row) for row in suppliers_raw]
    return render_template('supplier_management.html', suppliers=suppliers)


@app.route('/suppliers/add', methods=['POST'])
@login_required
def add_supplier():
    try:
        supplier_name = request.form['supplier_name']
        contact_number = request.form['contact_number']
        email = request.form.get('email', '')
        address = request.form.get('address', '')

        conn = get_db()
        conn.execute(
            'INSERT INTO suppliers (supplier_name, contact_number, email, address) '
            'VALUES (?, ?, ?, ?)',
            (supplier_name, contact_number, email, address)
        )
        conn.commit()
        conn.close()

        flash('Supplier added successfully!', 'success')
    except Exception as e:
        flash(f'Error adding supplier: {str(e)}', 'error')

    return redirect(url_for('suppliers'))


@app.route('/suppliers/update/<int:id>', methods=['POST'])
@login_required
def update_supplier(id):
    try:
        supplier_name = request.form['supplier_name']
        contact_number = request.form['contact_number']
        email = request.form.get('email', '')
        address = request.form.get('address', '')

        conn = get_db()
        conn.execute(
            'UPDATE suppliers '
            'SET supplier_name=?, phone=?, email=?, address=? '
            'WHERE supplier_id=?',
            (supplier_name, contact_number, email, address, id)
        )
        conn.commit()
        conn.close()

        flash('Supplier updated successfully!', 'success')
    except Exception as e:
        flash(f'Error updating supplier: {str(e)}', 'error')

    return redirect(url_for('suppliers'))


@app.route('/suppliers/delete/<int:id>')
@login_required
def delete_supplier(id):
    try:
        conn = get_db()
        conn.execute('DELETE FROM suppliers WHERE supplier_id = ?', (id,))
        conn.commit()
        conn.close()
        flash('Supplier deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting supplier: {str(e)}', 'error')

    return redirect(url_for('suppliers'))
@app.route('/reports')
@login_required
def reports():
    conn = get_db()
    today = date.today()

    # ---------- SALES SUMMARY ----------

    # Today sales
    today_sales_rows = conn.execute('''
        SELECT s.*, COUNT(si.item_id) AS items_count
        FROM sales s
        LEFT JOIN sale_items si ON s.sale_id = si.sale_id
        WHERE DATE(s.sale_date) = ?
        GROUP BY s.sale_id
        ORDER BY s.sale_date DESC
    ''', (today,)).fetchall()

    # Monthly sales (current month)
    first_day = today.replace(day=1)
    monthly_sales_rows = conn.execute('''
        SELECT s.*, COUNT(si.item_id) AS items_count
        FROM sales s
        LEFT JOIN sale_items si ON s.sale_id = si.sale_id
        WHERE DATE(s.sale_date) >= ?
        GROUP BY s.sale_id
        ORDER BY s.sale_date DESC
    ''', (first_day,)).fetchall()

    # Top selling medicines
    top_medicines_rows = conn.execute('''
        SELECT m.medicine_name,
               m.category,
               SUM(si.quantity) AS total_sold,
               SUM(si.subtotal) AS total_revenue
        FROM sale_items si
        JOIN medicines m ON si.medicine_id = m.medicine_id
        GROUP BY si.medicine_id
        ORDER BY total_sold DESC
        LIMIT 10
    ''').fetchall()

    # Least selling medicines
    least_medicines_rows = conn.execute('''
        SELECT m.medicine_name,
               m.category,
               SUM(si.quantity) AS total_sold,
               SUM(si.subtotal) AS total_revenue
        FROM sale_items si
        JOIN medicines m ON si.medicine_id = m.medicine_id
        GROUP BY si.medicine_id
        ORDER BY total_sold ASC
        LIMIT 10
    ''').fetchall()

    # ---------- STOCK REPORTS ----------

    # Low stock (0 < qty < 10)
    low_stock_rows = conn.execute('''
        SELECT m.*, s.supplier_name
        FROM medicines m
        LEFT JOIN suppliers s ON m.supplier_id = s.supplier_id
        WHERE m.quantity > 0 AND m.quantity < 10
        ORDER BY m.quantity ASC
    ''').fetchall()

    # Out of stock
    out_stock_rows = conn.execute('''
        SELECT m.*, s.supplier_name
        FROM medicines m
        LEFT JOIN suppliers s ON m.supplier_id = s.supplier_id
        WHERE m.quantity = 0
        ORDER BY m.medicine_name
    ''').fetchall()

    # Stock total value
    stock_value = conn.execute('''
        SELECT SUM(quantity * price) AS total_value,
               COUNT(*)                AS total_medicines,
               SUM(quantity)           AS total_quantity
        FROM medicines
    ''').fetchone()

    # ---------- EXPIRY REPORTS ----------

    expired_rows = conn.execute('''
        SELECT m.*, s.supplier_name
        FROM medicines m
        LEFT JOIN suppliers s ON m.supplier_id = s.supplier_id
        WHERE DATE(m.expiry_date) < DATE('now')
        ORDER BY m.expiry_date ASC
    ''').fetchall()

    near_expiry_rows = conn.execute('''
        SELECT m.*, s.supplier_name
        FROM medicines m
        LEFT JOIN suppliers s ON m.supplier_id = s.supplier_id
        WHERE DATE(m.expiry_date) BETWEEN DATE('now') AND DATE('now', '+30 days')
        ORDER BY m.expiry_date ASC
    ''').fetchall()

    # ---------- SUPPLIER REPORTS ----------

    supplier_sales_rows = conn.execute('''
        SELECT s.supplier_name,
               COUNT(DISTINCT si.sale_id) AS total_sales,
               SUM(si.quantity)           AS total_quantity,
               SUM(si.subtotal)           AS total_revenue
        FROM suppliers s
        JOIN medicines m ON s.supplier_id = m.supplier_id
        JOIN sale_items si ON m.medicine_id = si.medicine_id
        GROUP BY s.supplier_id
        ORDER BY total_revenue DESC
    ''').fetchall()

    conn.close()

    # Convert rows → dict
    today_sales      = [dict(r) for r in today_sales_rows]
    monthly_sales    = [dict(r) for r in monthly_sales_rows]
    top_medicines    = [dict(r) for r in top_medicines_rows]
    least_medicines  = [dict(r) for r in least_medicines_rows]
    low_stock        = [dict(r) for r in low_stock_rows]
    out_stock        = [dict(r) for r in out_stock_rows]
    expired          = [dict(r) for r in expired_rows]
    near_expiry      = [dict(r) for r in near_expiry_rows]
    supplier_sales   = [dict(r) for r in supplier_sales_rows]

    return render_template(
        'reports.html',
        today_sales=today_sales,
        monthly_sales=monthly_sales,
        top_medicines=top_medicines,
        least_medicines=least_medicines,
        low_stock=low_stock,
        out_stock=out_stock,
        stock_value=stock_value,
        expired=expired,
        near_expiry=near_expiry,
        supplier_sales=supplier_sales
    )
@app.route('/export-pdf')
@login_required
def export_pdf():
    try:
        conn = get_db()
        today = date.today()

        # --- get some simple data (today sales) ---
        sales = [dict(r) for r in conn.execute('''
            SELECT s.*, COUNT(si.item_id) AS items_count
            FROM sales s
            LEFT JOIN sale_items si ON s.sale_id = si.sale_id
            WHERE DATE(s.sale_date) = ?
            GROUP BY s.sale_id
            ORDER BY s.sale_date DESC
        ''', (today,)).fetchall()]

        conn.close()

        # --- create PDF in memory ---
        buffer = BytesIO()          # <<==== IMPORTANT
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        title = Paragraph(f"PHARMACY REPORTS - {today}", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))

        data = [['Sale ID', 'Customer', 'Items', 'Amount']]
        for s in sales[:20]:
            data.append([
                s['sale_id'],
                s.get('customer_name', 'Walk-in'),
                s['items_count'],
                f"₹{s['total_amount']:.2f}"
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'report_{today}.pdf'
        )
    except Exception as e:
        flash(f'PDF Error: {str(e)}', 'error')
        return redirect(url_for('reports'))
@app.route('/export-excel')
@login_required
def export_excel():
    try:
        conn = get_db()
        today = date.today()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Medicines"

        medicines = [dict(r) for r in conn.execute('''
            SELECT m.*, COALESCE(s.supplier_name, 'N/A') AS supplier_name
            FROM medicines m
            LEFT JOIN suppliers s ON m.supplier_id = s.supplier_id
            ORDER BY m.medicine_name
        ''').fetchall()]

        ws['A1'] = f'Pharmacy Inventory - {today}'
        ws['A1'].font = Font(bold=True, size=16)

        headers = ['Medicine', 'Category', 'Batch', 'Price', 'Quantity', 'Expiry', 'Supplier']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF")

        for r, m in enumerate(medicines, 4):
            ws.cell(r, 1, m['medicine_name'])
            ws.cell(r, 2, m['category'])
            ws.cell(r, 3, m['batch_number'])
            ws.cell(r, 4, m['price'])
            ws.cell(r, 5, m['quantity'])
            ws.cell(r, 6, m['expiry_date'])
            ws.cell(r, 7, m['supplier_name'])

        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 18

        buffer = BytesIO()          # <<==== IMPORTANT
        wb.save(buffer)
        buffer.seek(0)

        conn.close()

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reports_{today}.xlsx'
        )
    except Exception as e:
        flash(f'Excel Error: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/api/medicine/<int:id>')
@login_required
def get_medicine_api(id):
    conn = get_db()
    medicine = conn.execute('SELECT * FROM medicines WHERE medicine_id = ?', (id,)).fetchone()
    conn.close()
    
    if medicine:
        return jsonify(dict(medicine))
    return jsonify({'error': 'Medicine not found'}), 404
@app.route('/billing/invoice/<int:sale_id>')
@login_required
def view_invoice(sale_id):
    conn = get_db()
    
    # Get sale details
    sale = conn.execute('SELECT * FROM sales WHERE sale_id = ?', (sale_id,)).fetchone()
    
    if not sale:
        flash('Invoice not found', 'error')
        conn.close()
        return redirect(url_for('billing'))
    
    # Get sale items with medicine details
    items = conn.execute('''
        SELECT si.*, m.medicine_name, m.category, m.batch_number
        FROM sale_items si
        JOIN medicines m ON si.medicine_id = m.medicine_id
        WHERE si.sale_id = ?
    ''', (sale_id,)).fetchall()
    
    conn.close()
    
    return render_template('invoice.html', sale=dict(sale), items=[dict(item) for item in items])

# ==================== ADVANCED SUPPLIER MANAGEMENT ====================

@app.route('/suppliers/advanced')
@login_required
def suppliers_advanced():
    conn = get_db()
    
    # Get all suppliers with additional info
    suppliers = conn.execute('''
        SELECT s.*, 
               COUNT(DISTINCT po.po_id) as total_orders,
               COALESCE(SUM(po.total_amount), 0) as total_purchase,
               COALESCE(SUM(CASE WHEN sp.payment_id IS NULL THEN po.total_amount ELSE 0 END), 0) as outstanding
        FROM suppliers s
        LEFT JOIN purchase_orders po ON s.supplier_id = po.supplier_id
        LEFT JOIN supplier_payments sp ON po.po_id = sp.po_id
        GROUP BY s.supplier_id
        ORDER BY s.supplier_name
    ''').fetchall()
    
    conn.close()
    return render_template('suppliers_advanced.html', suppliers=[dict(s) for s in suppliers])
@app.route('/suppliers/profile/<int:supplier_id>')
@login_required
def supplier_profile(supplier_id):
    conn = get_db()
    
    # Get supplier details
    supplier = conn.execute('SELECT * FROM suppliers WHERE supplier_id = ?', (supplier_id,)).fetchone()
    
    if not supplier:
        flash('Supplier not found', 'error')
        conn.close()
        return redirect(url_for('suppliers_advanced'))
    
    # Initialize empty lists for now (tables will be populated as you use the system)
    orders = []
    payments = []
    returns = []
    
    # Try to get data if tables exist
    try:
        orders = conn.execute('''
            SELECT * FROM purchase_orders 
            WHERE supplier_id = ? 
            ORDER BY po_date DESC
        ''', (supplier_id,)).fetchall()
    except:
        pass
    
    try:
        payments = conn.execute('''
            SELECT * FROM supplier_payments 
            WHERE supplier_id = ? 
            ORDER BY payment_date DESC
        ''', (supplier_id,)).fetchall()
    except:
        pass
    
    try:
        returns = conn.execute('''
            SELECT pr.*, m.medicine_name 
            FROM purchase_returns pr
            JOIN medicines m ON pr.medicine_id = m.medicine_id
            WHERE pr.supplier_id = ?
            ORDER BY pr.return_date DESC
        ''', (supplier_id,)).fetchall()
    except:
        pass
    
    conn.close()
    
    return render_template('supplier_profile.html',
                         supplier=dict(supplier),
                         orders=[dict(o) for o in orders],
                         payments=[dict(p) for p in payments],
                         returns=[dict(r) for r in returns])

@app.route('/suppliers/add-advanced', methods=['POST'])
@login_required
def add_supplier_advanced():
    try:
        conn = get_db()
        conn.execute('''
            INSERT INTO suppliers 
            (supplier_name, phone, address, gstin, contact_person, 
             payment_terms, preferred_payment_mode, rating)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            request.form['supplier_name'],
            request.form['contact_number'],
            request.form.get('address', ''),
            request.form.get('gstin', ''),
            request.form.get('contact_person', ''),
            request.form.get('payment_terms', '30 Days'),
            request.form.get('preferred_payment_mode', 'Bank Transfer'),
            int(request.form.get('rating', 3))
        ))
        conn.commit()
        conn.close()
        
        flash('Supplier added successfully!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('suppliers_advanced'))

@app.route('/suppliers/update-advanced/<int:id>', methods=['POST'])
@login_required
def update_supplier_advanced(id):
    try:
        conn = get_db()
        conn.execute('''
            UPDATE suppliers 
            SET supplier_name=?, phone=?, address=?, 
                gstin=?, contact_person=?, payment_terms=?, 
                preferred_payment_mode=?, rating=?
            WHERE supplier_id=?
        ''', (
            request.form['supplier_name'],
            request.form['contact_number'],
            request.form.get('address', ''),
            request.form.get('gstin', ''),
            request.form.get('contact_person', ''),
            request.form.get('payment_terms', '30 Days'),
            request.form.get('preferred_payment_mode', 'Bank Transfer'),
            int(request.form.get('rating', 3)),
            id
        ))
        conn.commit()
        conn.close()
        
        flash('Supplier updated successfully!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('suppliers_advanced'))
# Purchase Order Routes
@app.route('/purchase-orders')
@login_required
def purchase_orders():
    conn = get_db()
    
    orders = conn.execute('''
        SELECT po.*, s.supplier_name 
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id = s.supplier_id
        ORDER BY po.po_date DESC
    ''').fetchall()
    
    suppliers = conn.execute('SELECT * FROM suppliers ORDER BY supplier_name').fetchall()
    
    conn.close()
    
    return render_template('purchase_orders.html', 
                         orders=[dict(o) for o in orders],
                         suppliers=[dict(s) for s in suppliers])

@app.route('/purchase-orders/add', methods=['POST'])
@login_required
def add_purchase_order():
    try:
        conn = get_db()
        
        # Create PO
        cursor = conn.execute('''
            INSERT INTO purchase_orders 
            (supplier_id, po_number, po_date, expected_delivery, total_amount, status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            request.form['supplier_id'],
            request.form['po_number'],
            request.form['po_date'],
            request.form.get('expected_delivery'),
            float(request.form['total_amount']),
            'Pending',
            request.form.get('notes', '')
        ))
        
        conn.commit()
        conn.close()
        
        flash('Purchase Order created successfully!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('purchase_orders'))



@app.route('/purchase-orders/receive/<int:po_id>', methods=['POST'])
@login_required
def receive_purchase_order(po_id):
    try:
        conn = get_db()
        
        # Update PO status
        conn.execute('UPDATE purchase_orders SET status = ? WHERE po_id = ?', ('Received', po_id))
        
        # Add medicines to inventory (implement based on your needs)
        
        conn.commit()
        conn.close()
        
        flash('Purchase Order marked as received!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('purchase_orders'))

# Supplier Payments
@app.route('/supplier-payments')
@login_required
def supplier_payments():
    conn = get_db()
    
    payments = conn.execute('''
        SELECT sp.*, s.supplier_name, po.po_number
        FROM supplier_payments sp
        JOIN suppliers s ON sp.supplier_id = s.supplier_id
        LEFT JOIN purchase_orders po ON sp.po_id = po.po_id
        ORDER BY sp.payment_date DESC
    ''').fetchall()
    
    # Outstanding amounts
    outstanding = conn.execute('''
        SELECT s.supplier_id, s.supplier_name, 
               COALESCE(SUM(po.total_amount), 0) - COALESCE(SUM(sp.amount), 0) as outstanding
        FROM suppliers s
        LEFT JOIN purchase_orders po ON s.supplier_id = po.supplier_id
        LEFT JOIN supplier_payments sp ON po.po_id = sp.po_id
        GROUP BY s.supplier_id
        HAVING outstanding > 0
    ''').fetchall()
    
    conn.close()
    
    return render_template('supplier_payments.html',
                         payments=[dict(p) for p in payments],
                         outstanding=[dict(o) for o in outstanding])

@app.route('/supplier-payments/add', methods=['POST'])
@login_required
def add_supplier_payment():
    try:
        conn = get_db()
        conn.execute('''
            INSERT INTO supplier_payments 
            (supplier_id, po_id, payment_date, amount, payment_mode, reference_number, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            request.form['supplier_id'],
            request.form.get('po_id'),
            request.form['payment_date'],
            float(request.form['amount']),
            request.form['payment_mode'],
            request.form.get('reference_number', ''),
            request.form.get('notes', '')
        ))
        conn.commit()
        conn.close()
        
        flash('Payment recorded successfully!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('supplier_payments'))

# Purchase Returns
@app.route('/purchase-returns')
@login_required
def purchase_returns():
    conn = get_db()
    
    returns = conn.execute('''
        SELECT pr.*, s.supplier_name, m.medicine_name
        FROM purchase_returns pr
        JOIN suppliers s ON pr.supplier_id = s.supplier_id
        JOIN medicines m ON pr.medicine_id = m.medicine_id
        ORDER BY pr.return_date DESC
    ''').fetchall()
    
    suppliers = conn.execute('SELECT * FROM suppliers').fetchall()
    medicines = conn.execute('SELECT * FROM medicines').fetchall()
    
    conn.close()
    
    return render_template('purchase_returns.html',
                         returns=[dict(r) for r in returns],
                         suppliers=[dict(s) for s in suppliers],
                         medicines=[dict(m) for m in medicines])



@app.route('/purchase-returns/add', methods=['POST'])
@login_required
def add_purchase_return():
    try:
        conn = get_db()
        conn.execute('''
            INSERT INTO purchase_returns 
            (supplier_id, medicine_id, return_date, quantity, reason, credit_amount, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            request.form['supplier_id'],
            request.form['medicine_id'],
            request.form['return_date'],
            int(request.form['quantity']),
            request.form['reason'],
            float(request.form['credit_amount']),
            request.form.get('notes', '')
        ))
        conn.commit()
        conn.close()
        
        flash('Return recorded successfully!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('purchase_returns'))



if __name__ == '__main__':
    with app.app_context():
        init_db()
        add_sales_extra_columns()  # <-- new
    app.run(debug=True, port=5000)
